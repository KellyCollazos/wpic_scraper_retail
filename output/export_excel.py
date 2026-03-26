"""
output/export_excel.py
Reads master_enriched from Google Sheets and exports a formatted .xlsx file.
One worksheet per vendor, sorted by ICP score descending.
"""

from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from utils.helpers import get_logger, get_or_create_sheet, sheet_to_records

log = get_logger("export_excel")

# ── Styling constants ─────────────────────────────────────────────

HEADER_FILL = PatternFill("solid", fgColor="1A3C5E")      # dark navy
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
TIER_COLORS = {
    "A": "D4EDDA",   # soft green
    "B": "D1ECF1",   # soft blue
    "C": "FFF3CD",   # soft amber
    "D": "F8D7DA",   # soft red
}
THIN_BORDER = Border(
    left=Side(style="thin", color="DDDDDD"),
    right=Side(style="thin", color="DDDDDD"),
    top=Side(style="thin", color="DDDDDD"),
    bottom=Side(style="thin", color="DDDDDD"),
)

DISPLAY_COLUMNS = [
    ("company_name",        "Company",          30),
    ("category",            "Category",         20),
    ("website",             "Website",          28),
    ("hq_country",          "HQ",               8),
    ("revenue_est",         "Revenue Est.",     16),
    ("icp_score",           "ICP Score",        10),
    ("icp_tier",            "Tier",             8),
    ("icp_reason",          "ICP Reason",       40),
    ("china_presence_flag", "China Presence",   14),
    ("listing_url",         "Listing URL",      35),
    ("date_scraped",        "Scraped",          18),
    ("enriched_at",         "Enriched",         18),
]


def export_to_excel(spreadsheet, cfg: dict):
    output_path = cfg["output"]["excel_path"]
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)

    enriched_sheet = cfg["google_sheets"]["enriched_sheet"]
    ws_source = get_or_create_sheet(spreadsheet, enriched_sheet)
    all_records = sheet_to_records(ws_source)

    if not all_records:
        log.warning("master_enriched is empty — nothing to export.")
        return

    # Group records by vendor
    by_vendor: dict[str, list[dict]] = {}
    for row in all_records:
        vendor = row.get("vendor", "Unknown")
        by_vendor.setdefault(vendor, []).append(row)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default blank sheet

    for vendor_name, records in sorted(by_vendor.items()):
        # Sort by ICP score descending
        records.sort(key=lambda r: int(r.get("icp_score") or 0), reverse=True)
        ws = wb.create_sheet(title=vendor_name[:31])  # Excel tab limit = 31 chars
        _write_sheet(ws, records)
        log.info(f"  Sheet '{vendor_name}': {len(records)} rows")

    # Summary sheet
    _write_summary_sheet(wb, by_vendor)

    wb.save(output_path)
    log.info(f"Saved: {output_path}")


def _write_sheet(ws, records: list[dict]):
    col_keys   = [c[0] for c in DISPLAY_COLUMNS]
    col_labels = [c[1] for c in DISPLAY_COLUMNS]
    col_widths = [c[2] for c in DISPLAY_COLUMNS]

    # Header row
    for col_idx, label in enumerate(col_labels, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = THIN_BORDER
    ws.row_dimensions[1].height = 20

    # Data rows
    for row_idx, record in enumerate(records, start=2):
        tier = str(record.get("icp_tier", ""))
        fill_color = TIER_COLORS.get(tier, "FFFFFF")
        row_fill = PatternFill("solid", fgColor=fill_color)

        for col_idx, key in enumerate(col_keys, start=1):
            value = record.get(key, "")
            if key == "china_presence_flag":
                value = "Yes" if str(value).lower() in ("true", "1", "yes") else "No"
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = row_fill
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=(key == "icp_reason"))

    # Column widths
    for col_idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Freeze header row
    ws.freeze_panes = "A2"

    # Auto-filter
    ws.auto_filter.ref = ws.dimensions


def _write_summary_sheet(wb, by_vendor: dict[str, list[dict]]):
    ws = wb.create_sheet(title="Summary", index=0)

    headers = ["Vendor", "Total Companies", "Tier A", "Tier B", "Tier C", "Tier D", "Avg ICP Score"]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="left")

    for row_idx, (vendor, records) in enumerate(sorted(by_vendor.items()), start=2):
        tiers = [str(r.get("icp_tier", "")) for r in records]
        scores = [int(r.get("icp_score") or 0) for r in records]
        avg = round(sum(scores) / len(scores), 1) if scores else 0

        ws.cell(row=row_idx, column=1, value=vendor)
        ws.cell(row=row_idx, column=2, value=len(records))
        ws.cell(row=row_idx, column=3, value=tiers.count("A"))
        ws.cell(row=row_idx, column=4, value=tiers.count("B"))
        ws.cell(row=row_idx, column=5, value=tiers.count("C"))
        ws.cell(row=row_idx, column=6, value=tiers.count("D"))
        ws.cell(row=row_idx, column=7, value=avg)

    for col_idx, width in enumerate([20, 18, 10, 10, 10, 10, 16], start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
